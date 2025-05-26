import os
import uuid
import subprocess
import logging
from flask import Flask, request, send_file, jsonify, after_this_request
from werkzeug.utils import secure_filename
from flask_cors import CORS
import openpyxl
import pdfkit

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'output'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def is_complex_excel(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet in wb.worksheets:
            if sheet.merged_cells.ranges or sheet.max_column > 10 or sheet.max_row > 50:
                return True
        return False
    except Exception as e:
        logger.warning(f"Excel complexity check failed: {e}")
        return True

def get_cell_style(cell):
    style = "padding: 5px; border: 1px solid #999;"
    fill = cell.fill
    if fill and fill.fgColor and fill.fgColor.type == "rgb" and fill.fgColor.rgb != '00000000':
        style += f"background-color: #{fill.fgColor.rgb[-6:]};"
    font = cell.font
    if font:
        if font.bold: style += "font-weight: bold;"
        if font.italic: style += "font-style: italic;"
        if font.color and font.color.type == "rgb":
            style += f"color: #{font.color.rgb[-6:]};"
        if font.size: style += f"font-size: {font.size}px;"
    alignment = cell.alignment
    if alignment:
        if alignment.horizontal: style += f"text-align: {alignment.horizontal};"
        if alignment.vertical: style += f"vertical-align: {alignment.vertical};"
    return style

def convert_excel_to_html(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.active
    html = """
    <html><head><meta charset="utf-8"/>
    <style>
    table { border-collapse: collapse; width: 100%; table-layout: fixed; font-family: Arial, sans-serif; }
    td, th { word-wrap: break-word; vertical-align: top; }
    </style></head><body><table>
    """
    for row in sheet.iter_rows():
        html += "<tr>"
        for cell in row:
            value = cell.value if cell.value is not None else ""
            style = get_cell_style(cell)
            html += f"<td style='{style}'>{value}</td>"
        html += "</tr>"
    html += "</table></body></html>"
    return html

def convert_using_libreoffice(input_path, output_dir):
    try:
        subprocess.run([
            'libreoffice', '--headless', '--convert-to', 'pdf', input_path, '--outdir', output_dir
        ], check=True)
        base_name = os.path.splitext(os.path.basename(input_path))[0]
        return os.path.join(output_dir, base_name + '.pdf')
    except subprocess.CalledProcessError as e:
        logger.error(f"LibreOffice failed: {e}")
        raise

def convert_using_html_to_pdf(input_path, output_pdf_path):
    html = convert_excel_to_html(input_path)
    html_path = output_pdf_path.replace('.pdf', '.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html)

    config = pdfkit.configuration(wkhtmltopdf=os.getenv("WKHTMLTOPDF_PATH", "/usr/bin/wkhtmltopdf"))
    pdfkit.from_file(html_path, output_pdf_path, configuration=config, options={
        'page-size': 'A4',
        'encoding': 'UTF-8',
        'margin-top': '10mm',
        'margin-bottom': '10mm',
        'margin-left': '10mm',
        'margin-right': '10mm',
    })

    os.remove(html_path)
    return output_pdf_path

@app.route('/convert', methods=['POST'])
def convert_excel_to_pdf():
    file = request.files.get('file')
    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Invalid file type. Only .xls or .xlsx allowed."}), 400

    filename = secure_filename(file.filename)
    unique_id = str(uuid.uuid4())
    input_path = os.path.join(UPLOAD_FOLDER, f"{unique_id}_{filename}")
    output_pdf_name = f"{os.path.splitext(filename)[0]}_{unique_id}.pdf"
    output_path = os.path.join(OUTPUT_FOLDER, output_pdf_name)

    file.save(input_path)

    try:
        if is_complex_excel(input_path):
            final_pdf = convert_using_libreoffice(input_path, OUTPUT_FOLDER)
        else:
            final_pdf = convert_using_html_to_pdf(input_path, output_path)

        @after_this_request
        def cleanup(response):
            paths_to_delete = [input_path, final_pdf]
            for path in paths_to_delete:
                try:
                    if os.path.exists(path):
                        os.remove(path)
                        logger.info(f"Deleted temporary file: {path}")
                except Exception as cleanup_error:
                    logger.warning(f"Cleanup failed for {path}: {cleanup_error}")
            return response

        return send_file(final_pdf, as_attachment=True, download_name=os.path.basename(final_pdf))

    except Exception as e:
        logger.exception("Conversion failed")
        try:
            if os.path.exists(input_path):
                os.remove(input_path)
        except:
            pass
        return jsonify({"error": "Conversion failed", "details": str(e)}), 500

if __name__ == '__main__':
    import os
    port = int(os.environ.get("PORT", 7000))
    app.run(host='0.0.0.0', port=port)
